#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Project  :interface_test
# @Description: 
# @Time    : 2022-01-05 21:40
# @Author  : Con 
# @Email    :meetky@sina.cn
# @Site    : 
# @File    : handle_data.py
# @Software: PyCharm

import openpyxl


class HandleExcel:
    """"""

    def __init__(self, file_name, sheet_name):
        self.file = file_name
        self.sheet = sheet_name

    def read_execl(self):
        # 读取数据
        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook[self.sheet]
        all_rows = list(sheet.rows)
        # 获取第一行
        title = [i.value for i in all_rows[0]]
        cases = []
        # 获取所有行，每行即是一条case
        for item in all_rows[1:]:
            # 获取case中的数据
            data = [i.value for i in item]
            case = dict(zip(title, data))
            cases.append(case)
        return cases

    def write_execl(self, row, column, value):
        # 加载工作簿
        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook[self.sheet]
        # 写入数据
        sheet.cell(row=row, column=column, value=value)
        workbook.save(self.file)


class Handle_Yaml:
    def __init__(self):
        pass

    def read_yaml(self):
        pass

    def write_yaml(self):
        pass


if __name__ == '__main__':
    he = HandleExcel(r"D:\Env_Project\interface-testframwork\datas\test.xlsx", "addUser")
    cases = he.read_execl()
    for case in cases:
        print("---------------------------")
        print(eval(case["expected"])["code"])
        print("---------------------------")
